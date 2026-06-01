from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse, JsonResponse
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.decorators import permission_required
from openpyxl import Workbook
from datetime import date
import pandas as pd
from rest_framework import request

from .models import Employee, EmployeeDocument
from .forms import EmployeeForm
from django.shortcuts import render, get_object_or_404 
from .models import Employee
from django.template.loader import get_template
from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse, JsonResponse
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.template.loader import get_template
from django.urls import reverse
import qrcode
from io import BytesIO
from django.core.files import File
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.drawing.image import Image as XLImage
from django.conf import settings
from django.db.models import Count
import os
import base64
import cv2
import numpy as np


EXCEL_COLUMN_ALIASES = {
    'employeeid': 'employee_id',
    'empid': 'employee_id',
    'name': 'name',
    'gender': 'gender',
    'fatherspousename': 'father_spouse_name',
    'fathername': 'father_spouse_name',
    'spousename': 'father_spouse_name',
    'dob': 'dob',
    'dateofbirth': 'dob',
    'placeofbirth': 'place_of_birth',
    'nationality': 'nationality',
    'education': 'education_level',
    'educationlevel': 'education_level',
    'joiningdate': 'joining_date',
    'department': 'department',
    'designation': 'designation',
    'category': 'category',
    'employmenttype': 'employment_type',
    'mobile': 'mobile',
    'mobileno': 'mobile',
    'mobilenumber': 'mobile',
    'uan': 'uan',
    'pan': 'pan',
    'nomineename': 'nominee_name',
    'epsnps': 'eps_nps',
    'familydetails': 'family_details',
    'postingdetails': 'posting_details',
    'pay': 'pay',
    'promotion': 'promotion',
    'esicip': 'esic_ip',
    'aadhaar': 'aadhaar',
    'aadhar': 'aadhaar',
    'bankaccount': 'bank_account_no',
    'bankaccountno': 'bank_account_no',
    'bankname': 'bank_name',
    'ifsc': 'ifsc',
    'presentaddress': 'present_address',
    'permanentaddress': 'permanent_address',
    'servicebookno': 'service_book_no',
    'exitdate': 'exit_date',
    'exitreason': 'exit_reason',
    'identificationmark': 'identification_mark',
    'remarks': 'remarks',
}


def normalize_excel_header(value):
    return ''.join(
        char for char in str(value).strip().lower()
        if char.isalnum()
    )


def clean_excel_value(value, default=''):
    if pd.isna(value):
        return default

    if isinstance(value, float) and value.is_integer():
        return str(int(value))

    return str(value).strip()


def clean_excel_date(value):
    if pd.isna(value) or value == '':
        return None

    if hasattr(value, 'date'):
        return value.date()

    return value


def first_allowed_url(user):
    permission_url_pairs = (
        ('employees.view_dashboard', 'dashboard'),
        ('employees.view_employee', 'employee_list'),
        ('employees.add_employee', 'add_employee'),
        ('employees.import_employee_excel', 'upload_excel_page'),
        ('employees.scan_employee_qr', 'qr_scanner'),
        ('employees.export_employee_excel', 'reports'),
    )

    for permission, url_name in permission_url_pairs:
        if user.has_perm(permission):
            return reverse(url_name)

    if user.is_staff:
        return '/admin/'

    return reverse('access_denied')
# LOGIN

def login_view(request):

    if request.user.is_authenticated:
        return redirect(first_allowed_url(request.user))

    error = ''

    if request.method == 'POST':

        username = request.POST.get('username')
        password = request.POST.get('password')

        user = authenticate(
            request,
            username=username,
            password=password
        )

        if user is not None and user.is_active:

            login(request, user)

            return redirect(first_allowed_url(user))

        else:

            error = 'Invalid username or password'

    return render(request, 'login.html', {
        'error': error
    })


def access_denied(request):
    return render(request, 'access_denied.html', status=403)


# LOGOUT

def logout_view(request):

    logout(request)

    return redirect('login')


# DASHBOARD

@login_required(login_url='login')
@permission_required('employees.view_dashboard', raise_exception=True)
def dashboard(request):

    employees = Employee.objects.all().order_by('-id')

    total_employees = Employee.objects.count()

    departments_count = Employee.objects.values(
        'department'
    ).distinct().count()

    documents_count = EmployeeDocument.objects.count()

    context = {

        'employees': employees,

        'total_employees': total_employees,

        'departments_count': departments_count,

        'documents_count': documents_count,

    }

    return render(
        request,
        'dashboard.html',
        context
    )


# EMPLOYEE LIST + SEARCH

@login_required(login_url='login')
@permission_required('employees.view_employee', raise_exception=True)
def employee_list(request):

    employees = Employee.objects.all().order_by('-id')

    today = date.today().isoformat()

    # SEARCH

    search = request.GET.get('search')

    if search:

        employees = (
            Employee.objects.filter(name__icontains=search) |
            Employee.objects.filter(employee_id__icontains=search) |
            Employee.objects.filter(department__icontains=search) |
            Employee.objects.filter(designation__icontains=search) |
            Employee.objects.filter(mobile__icontains=search)
        ).distinct().order_by('-id')

    # DATE FILTER

    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date', today)

    if start_date and end_date:

        employees = employees.filter(
            joining_date__range=[start_date, end_date]
        )

    return render(request, 'employees/employee_list.html', {

        'employees': employees,
        'today': today

    })


# ADD EMPLOYEE

@login_required(login_url='login')
@permission_required('employees.add_employee', raise_exception=True)
def add_employee(request):

    if request.method == 'POST':

        form = EmployeeForm(
            request.POST,
            request.FILES
        )

        if form.is_valid():

            employee = form.save()
            # GENERATE QR URL

            pdf_url = request.build_absolute_uri(
                reverse('employee_pdf', args=[employee.id])
            )

            # CREATE QR

            qr = qrcode.QRCode(
                version=1,
                box_size=10,
                border=5
            )

            qr.add_data(pdf_url)
            qr.make(fit=True)

            img = qr.make_image(fill_color="black", back_color="white")

            buffer = BytesIO()
            img.save(buffer, format='PNG')

            employee.qr_code.save(
                f'employee_{employee.id}.png',
                File(buffer),
                save=False
            )

            employee.save()

            # SAVE MULTIPLE DOCUMENTS

            files = request.FILES.getlist('documents')

            for file in files:

                EmployeeDocument.objects.create(
                    employee=employee,
                    document_name=file.name,
                    file=file
                )

            return redirect(
                'employee_detail',
                id=employee.id
            )

        else:

            print(form.errors)

    else:

        form = EmployeeForm()

    return render(
        request,
        'employees/employee_form.html',
        {
            'form': form
        }
    )


# EDIT EMPLOYEE

@login_required(login_url='login')
@permission_required('employees.change_employee', raise_exception=True)
def edit_employee(request, id):

    employee = get_object_or_404(Employee, id=id)

    if request.method == 'POST':

        form = EmployeeForm(
            request.POST,
            request.FILES,
            instance=employee
        )

        if form.is_valid():

            updated_employee = form.save()

            # SAVE NEW DOCUMENTS

            files = request.FILES.getlist('documents')

            for file in files:

                EmployeeDocument.objects.create(
                    employee=updated_employee,
                    document_name=file.name,
                    file=file
                )

            return redirect(
                'employee_detail',
                id=updated_employee.id
            )

        else:

            print(form.errors)

    else:

        form = EmployeeForm(instance=employee)

    return render(
        request,
        'employees/employee_form.html',
        {
            'form': form,
            'employee': employee,
            'edit_mode': True
        }
    )


# DELETE EMPLOYEE

@login_required(login_url='login')
@permission_required('employees.delete_employee', raise_exception=True)
def delete_employee(request, id):

    employee = get_object_or_404(Employee, id=id)

    # DELETE DOCUMENTS

    employee.documents.all().delete()

    # DELETE EMPLOYEE

    employee.delete()

    return redirect('employee_list')


# EMPLOYEE DETAIL

@login_required(login_url='login')
@permission_required('employees.view_employee', raise_exception=True)
def employee_detail(request, id):

    employee = get_object_or_404(
        Employee,
        id=id
    )

    return render(
        request,
        'employee_detail.html',
        {
            'employee': employee
        }
    )


@login_required(login_url='login')
@permission_required('employees.view_employee', raise_exception=True)
def employee_detail_by_code(request, employee_id):

    employee = get_object_or_404(
        Employee,
        employee_id=employee_id
    )

    return redirect(
        'employee_detail',
        id=employee.id
    )


def employee_pdf(request, id):

    employee = get_object_or_404(Employee, id=id)

    return render(
        request,
        'employees/employee_pdf.html',
        {
            'employee': employee
        }
    )


# EMPLOYEE API

def employee_api(request):

    employees = Employee.objects.all()

    data = []

    for employee in employees:

        documents = []

        for doc in employee.documents.all():

            documents.append({
                'document_name': doc.document_name,
                'file': request.build_absolute_uri(doc.file.url)
            })

        data.append({

            # BASIC DETAILS

            'id': employee.id,
            'employee_id': employee.employee_id,
            
            'name': employee.name,
            'gender': employee.gender,
            'father_spouse_name': employee.father_spouse_name,
            'dob': str(employee.dob) if employee.dob else '',
            'nationality': employee.nationality,
            'education_level': employee.education_level,
            'department': employee.department,
            'designation': employee.designation,
            'category': employee.category,
            'employment_type': employee.employment_type,
            'mobile': employee.mobile,
            'joining_date': str(employee.joining_date),


            # GOVERNMENT DETAILS

            'uan': employee.uan,
            'pan': employee.pan,
            'place_of_birth': employee.place_of_birth,

            'eps_nps': employee.eps_nps,

            'family_details': employee.family_details,

            'posting_details': employee.posting_details,

            'pay': employee.pay,

            'promotion': employee.promotion,
            'esic_ip': employee.esic_ip,
            'aadhaar': employee.aadhaar,

            # BANK DETAILS

            'bank_account_no': employee.bank_account_no,
            'bank_name': employee.bank_name,
            'ifsc': employee.ifsc,

            # ADDRESS

            'present_address': employee.present_address,
            'permanent_address': employee.permanent_address,

            # EXIT DETAILS

            'service_book_no': employee.service_book_no,
            'exit_date': str(employee.exit_date) if employee.exit_date else '',
            'exit_reason': employee.exit_reason,

            # IDENTIFICATION

            'identification_mark': employee.identification_mark,
            'remarks': employee.remarks,

         

           

            'nominee_name': employee.nominee_name,
            

            # IMAGES

            'photo': request.build_absolute_uri(employee.photo.url)
            if employee.photo else '',

            'signature': request.build_absolute_uri(employee.signature.url)
            if employee.signature else '',

            'qr_code': request.build_absolute_uri(employee.qr_code.url)
            if employee.qr_code else '',

            # DOCUMENTS

            'joining_letter': request.build_absolute_uri(employee.joining_letter.url)
            if employee.joining_letter else '',

            'appointment_letter': request.build_absolute_uri(employee.appointment_letter.url)
            if employee.appointment_letter else '',

            'documents': documents
            

        })

    return JsonResponse(data, safe=False)


# EMPLOYEE DETAIL API

def employee_detail_api(request, employee_id):

    employee = get_object_or_404(Employee, employee_id=employee_id)

    accept_header = request.headers.get('Accept', '')

    if 'text/html' in accept_header and 'application/json' not in accept_header:
        return redirect(
            'employee_detail',
            id=employee.id
        )

    documents = []

    for doc in employee.documents.all():

        documents.append({
            'document_name': doc.document_name,
            'file': request.build_absolute_uri(doc.file.url)
        })

    data = {

        'employee_id': employee.employee_id,
        'id': employee.id,
        
        'name': employee.name,
        'gender': employee.gender,
        'father_spouse_name': employee.father_spouse_name,
        'dob': str(employee.dob) if employee.dob else '',
        'nationality': employee.nationality,
        'education_level': employee.education_level,
        'department': employee.department,
        'designation': employee.designation,
        'category': employee.category,
        'employment_type': employee.employment_type,
        'mobile': employee.mobile,
        'joining_date': str(employee.joining_date),

        'uan': employee.uan,
        'pan': employee.pan,
        'place_of_birth': employee.place_of_birth,

        'eps_nps': employee.eps_nps,

        'family_details': employee.family_details,

        'posting_details': employee.posting_details,

        'pay': employee.pay,

        'promotion': employee.promotion,
        'esic_ip': employee.esic_ip,
        
        'aadhaar': employee.aadhaar,

        'bank_account_no': employee.bank_account_no,
        'bank_name': employee.bank_name,
        'ifsc': employee.ifsc,

        'present_address': employee.present_address,
        'permanent_address': employee.permanent_address,

        'service_book_no': employee.service_book_no,
        'exit_date': str(employee.exit_date) if employee.exit_date else '',
        'exit_reason': employee.exit_reason,

        'identification_mark': employee.identification_mark,
        'remarks': employee.remarks,

       

      

       

        'nominee_name': employee.nominee_name,
        

    

        'photo': request.build_absolute_uri(employee.photo.url)
        if employee.photo else '',

        'signature': request.build_absolute_uri(employee.signature.url)
        if employee.signature else '',

        'qr_code': request.build_absolute_uri(employee.qr_code.url)
        if employee.qr_code else '',

        'documents': documents

    }

    return JsonResponse(data)


# UPLOAD EXCEL

@login_required(login_url='login')
@permission_required('employees.import_employee_excel', raise_exception=True)
def upload_employee_excel(request):

    if request.method == "POST":

        excel_file = request.FILES.get("file")

        if not excel_file:

            return JsonResponse({
                "error": "No file uploaded"
            }, status=400)

        try:

            df = pd.read_excel(excel_file)
            df = df.rename(columns={
                column: EXCEL_COLUMN_ALIASES.get(
                    normalize_excel_header(column),
                    normalize_excel_header(column)
                )
                for column in df.columns
            })

            created = 0
            skipped = 0
            errors = []

            for index, row in df.iterrows():

                try:

                    row_number = index + 2
                    emp_id = clean_excel_value(row.get("employee_id", ""))

                    if not emp_id:

                        skipped += 1
                        errors.append(f"Row {row_number}: Employee ID missing")
                        continue

                    if Employee.objects.filter(
                        employee_id=emp_id
                    ).exists():

                        skipped += 1
                        errors.append(f"Row {row_number}: Employee ID {emp_id} already exists")
                        continue

                    required_fields = {
                        "Name": clean_excel_value(row.get("name", "")),
                        "Department": clean_excel_value(row.get("department", "")),
                        "Designation": clean_excel_value(row.get("designation", "")),
                        "Mobile": clean_excel_value(row.get("mobile", "")),
                        "Joining Date": clean_excel_date(row.get("joining_date", None)),
                    }

                    missing_fields = [
                        field_name
                        for field_name, field_value in required_fields.items()
                        if field_value in ("", None)
                    ]

                    if missing_fields:

                        skipped += 1
                        errors.append(
                            f"Row {row_number}: Missing {', '.join(missing_fields)}"
                        )
                        continue

                    employee = Employee(

                            employee_id=emp_id,
                            name=required_fields["Name"],
                            gender=clean_excel_value(row.get("gender", "")),
                            father_spouse_name=clean_excel_value(row.get("father_spouse_name", "")),

                            dob=clean_excel_date(row.get("dob", None)),
                            place_of_birth=clean_excel_value(row.get("place_of_birth", "")),

                            nationality=clean_excel_value(row.get("nationality", "")),
                            education_level=clean_excel_value(row.get("education_level", "")),

                            joining_date=required_fields["Joining Date"],

                            department=required_fields["Department"],
                            designation=required_fields["Designation"],
                            category=clean_excel_value(row.get("category", "")),
                            employment_type=clean_excel_value(row.get("employment_type", "")),

                            mobile=required_fields["Mobile"],

                            uan=clean_excel_value(row.get("uan", "")),
                            pan=clean_excel_value(row.get("pan", "")),

                            nominee_name=clean_excel_value(row.get("nominee_name", "")),

                            eps_nps=clean_excel_value(row.get("eps_nps", "")),
                            family_details=clean_excel_value(row.get("family_details", "")),
                            posting_details=clean_excel_value(row.get("posting_details", "")),

                            pay=clean_excel_value(row.get("pay", "")),
                            promotion=clean_excel_value(row.get("promotion", "")),

                            esic_ip=clean_excel_value(row.get("esic_ip", "")),
                            aadhaar=clean_excel_value(row.get("aadhaar", "")),

                            bank_account_no=clean_excel_value(row.get("bank_account_no", "")),
                            bank_name=clean_excel_value(row.get("bank_name", "")),
                            ifsc=clean_excel_value(row.get("ifsc", "")),

                            present_address=clean_excel_value(row.get("present_address", "")),
                            permanent_address=clean_excel_value(row.get("permanent_address", "")),

                            service_book_no=clean_excel_value(row.get("service_book_no", "")),

                            exit_date=clean_excel_date(row.get("exit_date", None)),
                            exit_reason=clean_excel_value(row.get("exit_reason", "")),

                            identification_mark=clean_excel_value(row.get("identification_mark", "")),
                            remarks=clean_excel_value(row.get("remarks", ""))
                        )
                    

                    employee.save()

                    # GENERATE QR CODE
                    pdf_url = request.build_absolute_uri(
                        reverse('employee_pdf', args=[employee.id])
                    )

                    qr = qrcode.QRCode(
                        version=1,
                        box_size=10,
                        border=5
                    )

                    qr.add_data(pdf_url)
                    qr.make(fit=True)

                    img = qr.make_image(
                        fill_color="black",
                        back_color="white"
                    )

                    buffer = BytesIO()
                    img.save(buffer, format='PNG')

                    employee.qr_code.save(
                        f'employee_{employee.id}.png',
                        File(buffer),
                        save=False
                    )

                    employee.save()

                    created += 1

                except Exception as e:

                    skipped += 1
                    errors.append(f"Row {index + 2}: {e}")

            return JsonResponse({

                "message": "Upload completed",
                "created": created,
                "skipped": skipped,
                "errors": errors[:20]

            })

        except Exception as e:

            return JsonResponse({

                "error": "Invalid Excel file",
                "details": str(e)

            }, status=400)

    return JsonResponse({
        "error": "Invalid request"
    }, status=400)


# EXPORT EXCEL


@login_required(login_url='login')
@permission_required('employees.export_employee_excel', raise_exception=True)
def reports(request):
    employees = Employee.objects.all().prefetch_related('documents').order_by('-id')

    today = date.today().isoformat()
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    department = request.GET.get('department')

    if start_date:
        employees = employees.filter(joining_date__gte=start_date)

    if end_date:
        employees = employees.filter(joining_date__lte=end_date)

    if department:
        employees = employees.filter(department=department)

    employee_ids = employees.values('id')
    total_employees = employees.count()
    documents_count = EmployeeDocument.objects.filter(employee_id__in=employee_ids).count()
    qr_generated_count = employees.exclude(qr_code='').exclude(qr_code__isnull=True).count()
    employees_with_documents = employees.filter(documents__isnull=False).distinct().count()
    missing_documents_count = total_employees - employees_with_documents

    department_breakdown = employees.values('department').annotate(
        total=Count('id')
    ).order_by('-total', 'department')

    category_breakdown = employees.exclude(category__isnull=True).exclude(
        category=''
    ).values('category').annotate(
        total=Count('id')
    ).order_by('-total', 'category')

    employment_breakdown = employees.exclude(employment_type__isnull=True).exclude(
        employment_type=''
    ).values('employment_type').annotate(
        total=Count('id')
    ).order_by('-total', 'employment_type')

    departments = Employee.objects.exclude(department='').values_list(
        'department',
        flat=True
    ).distinct().order_by('department')

    return render(request, 'reports.html', {
        'employees': employees[:10],
        'total_employees': total_employees,
        'documents_count': documents_count,
        'qr_generated_count': qr_generated_count,
        'missing_documents_count': missing_documents_count,
        'department_breakdown': department_breakdown,
        'category_breakdown': category_breakdown,
        'employment_breakdown': employment_breakdown,
        'departments': departments,
        'selected_department': department,
        'today': today,
    })


@login_required(login_url='login')
@permission_required('employees.export_employee_excel', raise_exception=True)
def export_employees_excel(request):

    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    department = request.GET.get('department')

    employees = Employee.objects.all()

    if start_date:
        employees = employees.filter(joining_date__gte=start_date)

    if end_date:
        employees = employees.filter(joining_date__lte=end_date)

    if department:
        employees = employees.filter(department=department)

    wb = Workbook()
    ws = wb.active
    ws.title = "Employee Register"


    headers = [

        "Employee ID",
        "Name",
        "Gender",
        "Father/Spouse Name",
        "DOB",
        "Place Of Birth",
        "Nationality",
        "Education Level",
        "Joining Date",
        "Department",
        "Designation",
        "Category",
        "Employment Type",
        "Mobile",
        "UAN",
        "PAN",
        "Nominee Name",
        "EPS/NPS",
        "Family Details",
        "Posting Details",
        "Pay",
        "Promotion",
        "ESIC IP",
        "AADHAAR",
        "Bank Account",
        "Bank Name",
        "IFSC",
        "Present Address",
        "Permanent Address",
        "Service Book No",
        "Exit Date",
        "Exit Reason",
        "Identification Mark",
        "Remarks",
        "Photo",
        "Signature",
        "QR Code"
    ]
    

    for col_num, header in enumerate(headers, 1):

        cell = ws.cell(row=1, column=col_num)
        cell.value = header

        cell.font = Font(
            bold=True,
            color="FFFFFF"
        )

        cell.fill = PatternFill(
            "solid",
            fgColor="1F4E78"
        )

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

    row_no = 2

    for emp in employees:

        ws.row_dimensions[row_no].height = 90

        values = [

               emp.employee_id,
                emp.name,
                emp.gender,
                emp.father_spouse_name,

                str(emp.dob) if emp.dob else "",
                emp.place_of_birth,

                emp.nationality,
                emp.education_level,

                str(emp.joining_date) if emp.joining_date else "",

                emp.department,
                emp.designation,
                emp.category,
                emp.employment_type,
                emp.mobile,

                emp.uan,
                emp.pan,

                emp.nominee_name,

                emp.eps_nps,
                emp.family_details,
                emp.posting_details,

                emp.pay,
                emp.promotion,

                emp.esic_ip,
                emp.aadhaar,

                emp.bank_account_no,
                emp.bank_name,
                emp.ifsc,

                emp.present_address,
                emp.permanent_address,

                emp.service_book_no,

                str(emp.exit_date) if emp.exit_date else "",
                emp.exit_reason,

                emp.identification_mark,
                emp.remarks,

                "",  # Photo
                "",  # Signature
                ""   # QR Code
        ]

        for col_num, value in enumerate(values, 1):
            ws.cell(
                row=row_no,
                column=col_num,
                value=value
            )

        photo_col = len(headers) - 2
        sign_col = len(headers) - 1
        qr_col = len(headers)

        try:
            if emp.photo:
                photo_path = emp.photo.path

                img = XLImage(photo_path)
                img.width = 70
                img.height = 70

                ws.add_image(
                    img,
                    f"AI{row_no}"
                )

        except:
            pass

        try:
            if emp.signature:
                sign_path = emp.signature.path

                img = XLImage(sign_path)
                img.width = 70
                img.height = 70

                ws.add_image(
                    img,
                    f"AJ{row_no}"
                )

        except:
            pass

        try:
            if emp.qr_code:
                qr_path = emp.qr_code.path

                img = XLImage(qr_path)
                img.width = 70
                img.height = 70

                ws.add_image(
                    img,
                    f"AK{row_no}"
                )

        except:
            pass

        row_no += 1

    for column in ws.columns:
        ws.column_dimensions[
            column[0].column_letter
        ].width = 25

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

    response['Content-Disposition'] = (
        'attachment; filename=Employee_Register.xlsx'
    )

    wb.save(response)

    return response
@login_required(login_url='login')
@permission_required('employees.import_employee_excel', raise_exception=True)
def upload_excel_page(request):
    return render(request, 'upload_excel_page.html')


@login_required(login_url='login')
@permission_required('employees.scan_employee_qr', raise_exception=True)
def qr_scanner(request):
    return render(request, 'qr_scanner.html')


@login_required(login_url='login')
@permission_required('employees.scan_employee_qr', raise_exception=True)
def decode_qr_code(request):

    if request.method != 'POST':
        return JsonResponse({
            'error': 'Invalid request'
        }, status=400)

    try:
        if request.FILES.get('image'):
            image_bytes = request.FILES['image'].read()
        else:
            image_data = request.POST.get('image', '')

            if ',' in image_data:
                image_data = image_data.split(',', 1)[1]

            image_bytes = base64.b64decode(image_data)

        image_array = np.frombuffer(image_bytes, dtype=np.uint8)
        image = cv2.imdecode(image_array, cv2.IMREAD_COLOR)

        if image is None:
            return JsonResponse({
                'error': 'Invalid image'
            }, status=400)

        detector = cv2.QRCodeDetector()
        value, _, _ = detector.detectAndDecode(image)

        if not value:
            return JsonResponse({
                'error': 'No QR code found'
            }, status=404)

        return JsonResponse({
            'value': value
        })

    except Exception as e:
        return JsonResponse({
            'error': str(e)
        }, status=400)
